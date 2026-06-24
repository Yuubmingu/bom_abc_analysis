import pandas as pd
import numpy as np
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# ============================================================
# 0. 기본 설정
# ============================================================

BASE_DIR = Path("/kaggle/working")

FILE_PATHS = {
    "production_plan": BASE_DIR / "production_plan.xlsx",
    "bom_master": BASE_DIR / "bom_master.xlsx",
    "material_master": BASE_DIR / "material_master.xlsx",
    "inventory": BASE_DIR / "inventory.xlsx",
}

OUTPUT_PATH = BASE_DIR / "bom_abc_analysis.xlsx"

REQUIRED_COLUMNS = {
    "production_plan": ["plan_week", "customer", "model", "product_code", "qty"],
    "bom_master": ["product_code", "material_code", "material_name", "usage_qty"],
    "material_master": ["material_code", "supplier", "lead_time_days", "moq", "unit_price"],
    "inventory": ["material_code", "current_stock"],
}

NUMERIC_COLUMNS = {
    "production_plan": ["qty"],
    "bom_master": ["usage_qty"],
    "material_master": ["lead_time_days", "moq", "unit_price"],
    "inventory": ["current_stock"],
}


# ============================================================
# 1. 공통 함수
# ============================================================

def check_file_exists(file_paths: dict):
    """
    입력 파일 존재 여부를 검증한다.
    파일이 없으면 어떤 파일이 없는지 명확하게 알려준다.
    """
    missing_files = []

    for name, path in file_paths.items():
        if not path.exists():
            missing_files.append(f"{name}: {path}")

    if missing_files:
        message = "\n".join(missing_files)
        raise FileNotFoundError(
            "[파일 존재 여부 검증 실패]\n"
            "아래 입력 파일이 존재하지 않습니다.\n\n"
            f"{message}"
        )


def read_excel_file(file_name: str, file_path: Path) -> pd.DataFrame:
    """
    Excel 파일을 읽는다.
    """
    try:
        df = pd.read_excel(file_path)
        print(f"[파일 읽기 완료] {file_name}: {file_path}")
        return df
    except Exception as e:
        raise RuntimeError(
            f"[파일 읽기 실패] {file_name}: {file_path}\n"
            f"오류 내용: {e}"
        )


def check_required_columns(df: pd.DataFrame, file_name: str, required_cols: list):
    """
    필수 컬럼 존재 여부를 검증한다.
    """
    missing_cols = [col for col in required_cols if col not in df.columns]

    if missing_cols:
        raise ValueError(
            f"[필수 컬럼 검증 실패] {file_name}\n"
            f"누락 컬럼: {missing_cols}\n"
            f"현재 컬럼: {list(df.columns)}"
        )


def convert_numeric_columns(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    """
    숫자 컬럼을 숫자형으로 변환한다.
    변환할 수 없는 값과 빈값은 0으로 처리한다.
    """
    df = df.copy()

    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def clean_text_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    주요 코드성 컬럼의 앞뒤 공백을 제거한다.
    Excel에서 코드 뒤에 공백이 들어간 경우 merge 오류를 줄이기 위함이다.
    """
    df = df.copy()

    text_cols = [
        "plan_week",
        "customer",
        "model",
        "product_code",
        "material_code",
        "material_name",
        "supplier",
    ]

    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    return df


def sort_plan_week_values(plan_weeks):
    """
    plan_week 값을 정렬한다.
    예: 2026-W27, 2026-W28 형태를 안정적으로 정렬한다.
    """
    def parse_week(x):
        try:
            text = str(x).strip()
            year_text, week_text = text.split("-W")
            return int(year_text), int(week_text)
        except Exception:
            return 9999, 9999

    return sorted(plan_weeks, key=parse_week)


def classify_abc(cumulative_ratio: float) -> str:
    """
    누적 비중 기준으로 ABC 등급을 분류한다.
    A: 80% 이하
    B: 80% 초과 ~ 95% 이하
    C: 95% 초과
    """
    if cumulative_ratio <= 0.80:
        return "A"
    elif cumulative_ratio <= 0.95:
        return "B"
    else:
        return "C"


def validate_calculation(df: pd.DataFrame):
    """
    핵심 계산식이 맞는지 검증한다.
    """
    print("\n==============================")
    print("계산 검증 시작")
    print("==============================")

    # 1) material_required_qty = qty × usage_qty 검증
    required_check = np.isclose(
        df["material_required_qty"],
        df["qty"] * df["usage_qty"],
        equal_nan=True
    ).all()

    if not required_check:
        raise ValueError("[계산 검증 실패] material_required_qty 계산이 맞지 않습니다.")

    print("[검증 완료] material_required_qty = qty × usage_qty")


def validate_final_calculation(abc_analysis: pd.DataFrame):
    """
    최종 ABC 분석 결과의 계산식을 검증한다.
    """
    # 2) usage_amount_12week = total_12week_usage_qty × unit_price 검증
    usage_amount_check = np.isclose(
        abc_analysis["usage_amount_12week"],
        abc_analysis["total_12week_usage_qty"] * abc_analysis["unit_price"],
        equal_nan=True
    ).all()

    if not usage_amount_check:
        raise ValueError("[계산 검증 실패] usage_amount_12week 계산이 맞지 않습니다.")

    print("[검증 완료] usage_amount_12week = total_12week_usage_qty × unit_price")

    # 3) cumulative_ratio = cumulative_amount / total_usage_amount 검증
    total_usage_amount = abc_analysis["usage_amount_12week"].sum()

    if total_usage_amount > 0:
        cumulative_ratio_check = np.isclose(
            abc_analysis["cumulative_ratio"],
            abc_analysis["cumulative_amount"] / total_usage_amount,
            equal_nan=True
        ).all()

        if not cumulative_ratio_check:
            raise ValueError("[계산 검증 실패] cumulative_ratio 계산이 맞지 않습니다.")

        print("[검증 완료] cumulative_ratio = cumulative_amount / total_usage_amount")
    else:
        print("[검증 참고] 전체 사용금액이 0이므로 cumulative_ratio 검증을 생략합니다.")

    # 4) ABC 등급 검증
    invalid_a = abc_analysis[
        (abc_analysis["abc_class"] == "A") &
        (abc_analysis["cumulative_ratio"] > 0.80)
    ]

    invalid_b = abc_analysis[
        (abc_analysis["abc_class"] == "B") &
        ~((abc_analysis["cumulative_ratio"] > 0.80) & (abc_analysis["cumulative_ratio"] <= 0.95))
    ]

    invalid_c = abc_analysis[
        (abc_analysis["abc_class"] == "C") &
        (abc_analysis["cumulative_ratio"] <= 0.95)
    ]

    if len(invalid_a) > 0 or len(invalid_b) > 0 or len(invalid_c) > 0:
        raise ValueError("[ABC 등급 검증 실패] ABC 분류 기준에 맞지 않는 행이 있습니다.")

    print("[검증 완료] ABC 등급 기준 검증 완료")


def adjust_excel_column_width(ws, min_width=10, max_width=35):
    """
    Excel 컬럼 너비를 데이터 길이에 맞게 자동 조정한다.
    너무 넓어지는 것을 방지하기 위해 최대 너비를 제한한다.
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, value_length)
            except Exception:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_common_sheet_style(ws):
    """
    모든 시트에 공통 적용할 기본 서식.
    """
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )

    # 첫 행 서식
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 전체 셀 기본 정렬과 테두리
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 필터 적용
    ws.auto_filter.ref = ws.dimensions

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 컬럼 너비 자동 조정
    adjust_excel_column_width(ws)


def apply_number_format(ws):
    """
    컬럼명 기준으로 숫자, 금액, 비율 서식을 적용한다.
    """
    header_map = {}

    for cell in ws[1]:
        header_map[cell.value] = cell.column

    amount_cols = [
        "unit_price",
        "usage_amount_12week",
        "total_usage_amount",
        "a_class_usage_amount",
        "usage_amount",
        "cumulative_amount",
        "amount_ratio",
    ]

    qty_cols = [
        "qty",
        "usage_qty",
        "material_required_qty",
        "required_qty",
        "total_12week_usage_qty",
        "total_usage_qty",
        "current_stock",
        "lead_time_days",
        "moq",
        "material_count",
        "a_class_count",
    ]

    ratio_cols = [
        "cumulative_ratio",
        "amount_ratio",
        "a_class_amount_ratio",
    ]

    for col_name, col_idx in header_map.items():
        col_letter = get_column_letter(col_idx)

        if col_name in ratio_cols:
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "0.0%"

        elif col_name in amount_cols:
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = '#,##0'

        elif col_name in qty_cols:
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = '#,##0'


def apply_abc_row_style(ws):
    """
    abc_class 기준으로 행 배경색을 적용한다.
    A: 주황/빨강 계열
    B: 노랑
    C: 연회색
    """
    header_map = {cell.value: cell.column for cell in ws[1]}

    if "abc_class" not in header_map:
        return

    abc_col_idx = header_map["abc_class"]

    fill_a = PatternFill("solid", fgColor="F4B183")  # 주황
    fill_b = PatternFill("solid", fgColor="FFF2CC")  # 노랑
    fill_c = PatternFill("solid", fgColor="E7E6E6")  # 연회색

    for row in range(2, ws.max_row + 1):
        abc_value = ws.cell(row=row, column=abc_col_idx).value

        if abc_value == "A":
            fill = fill_a
        elif abc_value == "B":
            fill = fill_b
        elif abc_value == "C":
            fill = fill_c
        else:
            fill = None

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def highlight_summary_a_class(ws):
    """
    summary_by_abc 시트에서 A등급 행을 강조한다.
    """
    header_map = {cell.value: cell.column for cell in ws[1]}

    if "abc_class" not in header_map:
        return

    abc_col_idx = header_map["abc_class"]
    fill_a = PatternFill("solid", fgColor="F4B183")
    bold_font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=abc_col_idx).value == "A":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_a
                ws.cell(row=row, column=col).font = bold_font


# ============================================================
# 2. 파일 존재 여부 검증 및 파일 읽기
# ============================================================

check_file_exists(FILE_PATHS)

production_plan = read_excel_file("production_plan", FILE_PATHS["production_plan"])
bom_master = read_excel_file("bom_master", FILE_PATHS["bom_master"])
material_master = read_excel_file("material_master", FILE_PATHS["material_master"])
inventory = read_excel_file("inventory", FILE_PATHS["inventory"])


# ============================================================
# 3. 필수 컬럼 검증
# ============================================================

check_required_columns(production_plan, "production_plan", REQUIRED_COLUMNS["production_plan"])
check_required_columns(bom_master, "bom_master", REQUIRED_COLUMNS["bom_master"])
check_required_columns(material_master, "material_master", REQUIRED_COLUMNS["material_master"])
check_required_columns(inventory, "inventory", REQUIRED_COLUMNS["inventory"])

print("\n[검증 완료] 모든 입력 파일의 필수 컬럼이 존재합니다.")


# ============================================================
# 4. 데이터 전처리
# ============================================================

production_plan = clean_text_columns(production_plan)
bom_master = clean_text_columns(bom_master)
material_master = clean_text_columns(material_master)
inventory = clean_text_columns(inventory)

production_plan = convert_numeric_columns(production_plan, NUMERIC_COLUMNS["production_plan"])
bom_master = convert_numeric_columns(bom_master, NUMERIC_COLUMNS["bom_master"])
material_master = convert_numeric_columns(material_master, NUMERIC_COLUMNS["material_master"])
inventory = convert_numeric_columns(inventory, NUMERIC_COLUMNS["inventory"])

material_master["supplier"] = material_master["supplier"].replace(["nan", "None", ""], np.nan).fillna("Unknown")


# ============================================================
# 5. 12주 대상 주차 선정
# ============================================================

all_plan_weeks = production_plan["plan_week"].dropna().unique().tolist()
sorted_plan_weeks = sort_plan_week_values(all_plan_weeks)
target_12_weeks = sorted_plan_weeks[:12]

production_plan_12week = production_plan[
    production_plan["plan_week"].isin(target_12_weeks)
].copy()

print("\n==============================")
print("12주 분석 대상 주차")
print("==============================")
print(target_12_weeks)


# ============================================================
# 6. BOM 매칭 검증
# ============================================================

plan_product_codes = set(production_plan_12week["product_code"].dropna().unique())
bom_product_codes = set(bom_master["product_code"].dropna().unique())

missing_bom_product_codes = sorted(list(plan_product_codes - bom_product_codes))

if missing_bom_product_codes:
    print("\n[경고] production_plan에는 있으나 bom_master에 없는 product_code가 있습니다.")
    print(missing_bom_product_codes)
else:
    print("\n[검증 완료] 모든 생산계획 product_code가 BOM에 존재합니다.")


# ============================================================
# 7. 생산계획과 BOM 연결
# ============================================================

plan_bom = production_plan_12week.merge(
    bom_master,
    on="product_code",
    how="left"
)

# BOM에 없는 제품은 자재 계산이 불가능하므로 제외
plan_bom = plan_bom.dropna(subset=["material_code"]).copy()

# 자재별 필요수량 계산
plan_bom["material_required_qty"] = plan_bom["qty"] * plan_bom["usage_qty"]

validate_calculation(plan_bom)


# ============================================================
# 8. 주차별 자재 사용량 집계
# ============================================================

weekly_usage = (
    plan_bom
    .groupby(["plan_week", "material_code", "material_name"], as_index=False)
    .agg(required_qty=("material_required_qty", "sum"))
)


# ============================================================
# 9. 자재별 12주 총 사용량 집계
# ============================================================

material_usage_12week = (
    weekly_usage
    .groupby(["material_code", "material_name"], as_index=False)
    .agg(total_12week_usage_qty=("required_qty", "sum"))
)


# ============================================================
# 10. 자재마스터 연결
# ============================================================

abc_analysis = material_usage_12week.merge(
    material_master[["material_code", "supplier", "lead_time_days", "moq", "unit_price"]],
    on="material_code",
    how="left"
)

abc_analysis["supplier"] = abc_analysis["supplier"].replace(["nan", "None", ""], np.nan).fillna("Unknown")
abc_analysis["lead_time_days"] = pd.to_numeric(abc_analysis["lead_time_days"], errors="coerce").fillna(0)
abc_analysis["moq"] = pd.to_numeric(abc_analysis["moq"], errors="coerce").fillna(0)
abc_analysis["unit_price"] = pd.to_numeric(abc_analysis["unit_price"], errors="coerce").fillna(0)


# ============================================================
# 11. 단가 누락 검증
# ============================================================

missing_price_items = abc_analysis[
    (abc_analysis["unit_price"].isna()) | (abc_analysis["unit_price"] == 0)
]["material_code"].dropna().unique().tolist()

if missing_price_items:
    print("\n[경고] unit_price가 없거나 0인 자재가 있습니다.")
    print(missing_price_items)
else:
    print("\n[검증 완료] 모든 사용 자재에 unit_price가 존재합니다.")


# ============================================================
# 12. 재고 정보 연결
# ============================================================

abc_analysis = abc_analysis.merge(
    inventory[["material_code", "current_stock"]],
    on="material_code",
    how="left"
)

abc_analysis["current_stock"] = pd.to_numeric(
    abc_analysis["current_stock"],
    errors="coerce"
).fillna(0)


# ============================================================
# 13. 12주 계획 사용금액 계산
# ============================================================

abc_analysis["usage_amount_12week"] = (
    abc_analysis["total_12week_usage_qty"] * abc_analysis["unit_price"]
)


# ============================================================
# 14. 사용금액 기준 정렬 및 누적 비중 계산
# ============================================================

abc_analysis = abc_analysis.sort_values(
    by="usage_amount_12week",
    ascending=False
).reset_index(drop=True)

total_usage_amount = abc_analysis["usage_amount_12week"].sum()

abc_analysis["amount_ratio"] = np.where(
    total_usage_amount > 0,
    abc_analysis["usage_amount_12week"] / total_usage_amount,
    0
)

abc_analysis["cumulative_amount"] = abc_analysis["usage_amount_12week"].cumsum()

abc_analysis["cumulative_ratio"] = np.where(
    total_usage_amount > 0,
    abc_analysis["cumulative_amount"] / total_usage_amount,
    0
)


# ============================================================
# 15. ABC 등급 분류
# ============================================================

abc_analysis["abc_class"] = abc_analysis["cumulative_ratio"].apply(classify_abc)


# 최종 컬럼 순서 정리
abc_analysis = abc_analysis[
    [
        "material_code",
        "material_name",
        "supplier",
        "lead_time_days",
        "moq",
        "unit_price",
        "current_stock",
        "total_12week_usage_qty",
        "usage_amount_12week",
        "amount_ratio",
        "cumulative_amount",
        "cumulative_ratio",
        "abc_class",
    ]
]

validate_final_calculation(abc_analysis)


# ============================================================
# 16. top_priority_items 생성_수정
# ============================================================



# ============================================================
# 17. summary_by_abc 생성
# ============================================================

summary_by_abc = (
    abc_analysis
    .groupby("abc_class", as_index=False)
    .agg(
        material_count=("material_code", "nunique"),
        total_usage_qty=("total_12week_usage_qty", "sum"),
        total_usage_amount=("usage_amount_12week", "sum"),
    )
)

summary_by_abc["amount_ratio"] = np.where(
    total_usage_amount > 0,
    summary_by_abc["total_usage_amount"] / total_usage_amount,
    0
)

# A, B, C 순서로 정렬
abc_order = pd.CategoricalDtype(categories=["A", "B", "C"], ordered=True)
summary_by_abc["abc_class"] = summary_by_abc["abc_class"].astype(abc_order)
summary_by_abc = summary_by_abc.sort_values("abc_class").reset_index(drop=True)
summary_by_abc["abc_class"] = summary_by_abc["abc_class"].astype(str)


# ============================================================
# 18. supplier_summary 생성
# ============================================================

supplier_summary = (
    abc_analysis
    .groupby("supplier", as_index=False)
    .agg(
        material_count=("material_code", "nunique"),
        total_usage_amount=("usage_amount_12week", "sum"),
    )
)

a_class_by_supplier = (
    abc_analysis[abc_analysis["abc_class"] == "A"]
    .groupby("supplier", as_index=False)
    .agg(
        a_class_count=("material_code", "nunique"),
        a_class_usage_amount=("usage_amount_12week", "sum"),
    )
)

supplier_summary = supplier_summary.merge(
    a_class_by_supplier,
    on="supplier",
    how="left"
)

supplier_summary["a_class_count"] = supplier_summary["a_class_count"].fillna(0).astype(int)
supplier_summary["a_class_usage_amount"] = supplier_summary["a_class_usage_amount"].fillna(0)

supplier_summary = supplier_summary.sort_values(
    by=["total_usage_amount", "a_class_count"],
    ascending=[False, False]
).reset_index(drop=True)

supplier_summary = supplier_summary[
    [
        "supplier",
        "material_count",
        "a_class_count",
        "total_usage_amount",
        "a_class_usage_amount",
    ]
]


# ============================================================
# 19. weekly_usage_detail 생성
# ============================================================

weekly_usage_detail = weekly_usage.merge(
    abc_analysis[
        [
            "material_code",
            "supplier",
            "unit_price",
            "abc_class",
        ]
    ],
    on="material_code",
    how="left"
)

weekly_usage_detail["supplier"] = weekly_usage_detail["supplier"].fillna("Unknown")
weekly_usage_detail["unit_price"] = pd.to_numeric(
    weekly_usage_detail["unit_price"],
    errors="coerce"
).fillna(0)

weekly_usage_detail["usage_amount"] = (
    weekly_usage_detail["required_qty"] * weekly_usage_detail["unit_price"]
)

weekly_usage_detail = weekly_usage_detail[
    [
        "plan_week",
        "material_code",
        "material_name",
        "required_qty",
        "unit_price",
        "usage_amount",
        "supplier",
        "abc_class",
    ]
]

weekly_usage_detail = weekly_usage_detail.sort_values(
    by=["plan_week", "usage_amount"],
    ascending=[True, False]
).reset_index(drop=True)


# ============================================================
# 20. Excel 파일 저장
# ============================================================

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    abc_analysis.to_excel(writer, sheet_name="abc_analysis", index=False)
    summary_by_abc.to_excel(writer, sheet_name="summary_by_abc", index=False)
    top_priority_items.to_excel(writer, sheet_name="top_priority_items", index=False)
    supplier_summary.to_excel(writer, sheet_name="supplier_summary", index=False)
    weekly_usage_detail.to_excel(writer, sheet_name="weekly_usage_detail", index=False)

    wb = writer.book

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        apply_common_sheet_style(ws)
        apply_number_format(ws)

        if sheet_name in ["abc_analysis", "top_priority_items", "weekly_usage_detail"]:
            apply_abc_row_style(ws)

        if sheet_name == "summary_by_abc":
            highlight_summary_a_class(ws)

    wb.save(OUTPUT_PATH)


# ============================================================
# 21. 완료 보고
# ============================================================

class_counts = abc_analysis["abc_class"].value_counts().to_dict()

a_class_amount = abc_analysis.loc[
    abc_analysis["abc_class"] == "A",
    "usage_amount_12week"
].sum()

a_class_amount_ratio = a_class_amount / total_usage_amount if total_usage_amount > 0 else 0

if len(abc_analysis) > 0:
    top_material_row = abc_analysis.iloc[0]
    top_material_text = (
        f"{top_material_row['material_code']} / "
        f"{top_material_row['material_name']} / "
        f"{top_material_row['usage_amount_12week']:,.0f}"
    )
else:
    top_material_text = "없음"

if len(supplier_summary) > 0:
    top_supplier_row = supplier_summary.iloc[0]
    top_supplier_text = (
        f"{top_supplier_row['supplier']} / "
        f"{top_supplier_row['total_usage_amount']:,.0f}"
    )
else:
    top_supplier_text = "없음"

print("\n==============================")
print("BOM ABC 분석 완료")
print("==============================")
print(f"생성 파일 경로: {OUTPUT_PATH}")
print(f"전체 자재 수: {abc_analysis['material_code'].nunique():,}")
print(f"A등급 자재 수: {class_counts.get('A', 0):,}")
print(f"B등급 자재 수: {class_counts.get('B', 0):,}")
print(f"C등급 자재 수: {class_counts.get('C', 0):,}")
print(f"전체 12주 사용금액: {total_usage_amount:,.0f}")
print(f"A등급 사용금액 합계: {a_class_amount:,.0f}")
print(f"A등급 사용금액 비중: {a_class_amount_ratio:.1%}")
print(f"사용금액 1위 자재: {top_material_text}")
print(f"사용금액 1위 협력사: {top_supplier_text}")
